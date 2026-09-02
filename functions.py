from time import strftime
from PySide6.QtCore import QDate, Qt
from PySide6.QtGui import QTextCharFormat, QColor, QBrush
from PySide6.QtWidgets import QFileDialog, QMessageBox, QTableWidgetItem
import json
from datetime import datetime, date
from collections import defaultdict
from services.plan_service import save_plan, load_plan, convert_amount
from services.records_service import load_records


def add_record(window):
    from ui.add_record_window import AddRecordWindow
    window.add_window = AddRecordWindow()
    window.add_window.show()

def show_summary(window):
    from ui.show_summary import ShowSummary
    window.add_window = ShowSummary()
    window.add_window.show()

def upload_new_plan(window):
    file_path, _ = QFileDialog.getOpenFileName(
        window, "Choose Excel file",
         "",
        "Excel files (*.xlsx)"

    )
    if not file_path:
        return

    try:
        from openpyxl import load_workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active
# we made data list
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            QMessageBox.warning(window, "Empty File", "Excel file is empty.")
            return

        headers = rows[0] # cia bus Device, data, amount
        data_rows = rows[1:] # cia bus FOD6020, 2026-01-05, 5. visos eilutes

        window.table.clear()
        window.table.setColumnCount(len(headers))
        window.table.setRowCount(len(data_rows))

        formatted_headers = []
        for header in headers:
            if isinstance(header, (datetime, date)):
                formatted_headers.append(header.strftime("%Y-%m-%d"))
            else:
                formatted_headers.append(str(header))
        window.table.setHorizontalHeaderLabels(formatted_headers)

        for row_index, row_data in enumerate(data_rows):
            for col_index, value in enumerate(row_data):
               # item = QTableWidgetItem("" if value is None else str(value))
                if value is None:
                    text = ""
                elif isinstance(value, datetime):
                    text = value.strftime("%d-%b")
                # its for numbers in plan is integer
                elif isinstance(value, float) and value.is_integer():
                    text = str(int(value))
                else:
                    text = str(value)

                item = QTableWidgetItem(text)
                item.setTextAlignment(Qt.AlignCenter)
                window.table.setItem(row_index, col_index, item)

        save_table_plan(window)

    except Exception as error:
        QMessageBox.critical(window, "Error", f"Could not load Excel file:\n{error}")

# Takes data from the table
def save_table_plan(window):
    headers = []
    for column in range(window.table.columnCount()):
        header_item = window.table.horizontalHeaderItem(column)
        headers.append(header_item.text())

    rows = []
    for row in range(window.table.rowCount()):
        row_data = []

        for column in range(window.table.columnCount()):
            item = window.table.item(row, column)

            if item is None:
                row_data.append("")
            else:
                row_data.append(item.text())

        rows.append(row_data)

    save_plan(headers, rows)

# Shows data from json to table
def load_saved_plan(window):
    plan = load_plan()

    if plan is None:
        return

    headers = plan["headers"]
    rows = plan["rows"]

    records = load_records()
    completed_amounts = defaultdict(int)

    for record in records:
        device = record.get("device")

        allocations = record.get("allocations", [])

        for allocation in allocations:
            target = allocation["target"]
            amount = convert_amount(allocation["amount"])

            key = (device, target)

            completed_amounts[key] += amount


    window.table.setColumnCount(len(headers))
    window.table.setRowCount(len(rows))
    window.table.setHorizontalHeaderLabels(headers)

    for row_index, row_data in enumerate(rows):
        device = row_data[0]

        for column_index, value in enumerate(row_data):
            if column_index == 0:
                text = str(value)

            elif value in ("", None):
                text = ""

            else:
                target = str(headers[column_index])
                planned_amount = convert_amount(value)
                key = (device, target)
                completed_amount = completed_amounts[key]
                text = (
                    f"{completed_amount} / "
                    f"{planned_amount}"
                )

            item = QTableWidgetItem(text)
            item.setTextAlignment(Qt.AlignCenter)

            if column_index > 0 and value not in ("", None):
                if completed_amount >= planned_amount:
                    item.setBackground(
                        QBrush(QColor("#C6EFCE"))
                    )
                else:
                    item.setBackground(
                        QBrush(QColor("#FFC7CE"))
                    )

            window.table.setItem(
                row_index,
                column_index,
                item
            )

def highlight_selected_device(window, selected_device):
    window.table.clearSelection()

    for row in range(window.table.rowCount()):
        item = window.table.item(row, 0)

        if item and item.text() == selected_device:
            window.table.selectRow(row)
            break


